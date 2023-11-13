import openpyxl
import asyncio

from aiogram import Router, F
from aiogram.types import Message
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext

from keyboards import reply, inline
from utils.states import Search
from config_reader import config

router = Router()
inv_book = openpyxl.load_workbook(filename = f"data/{config.namefile.get_secret_value()}.xlsx", data_only = True)

for item in inv_book.sheetnames:
    @router.message(F.text == item)
    async def comm_inventarization (message: Message):
        sheet = inv_book[message.text]
        text = ""
        for i in range(sheet.min_row, sheet.max_row + 1):
            text += f"🔢 № {sheet['A' + str(i)].value}\n📅 Дата ввода: {sheet['B' + str(i)].value}\n✏️ Сер. номер: {sheet['C' + str(i)].value}\n📌 Инв. номер: {sheet['D' + str(i)].value}\n📝 Наименование: {sheet['E' + str(i)].value}\n🚪 Местоположение: {sheet['F' + str(i)].value}\n\n"
        for x in range(0, len(text), 4096):
            mess = text[x: x + 4096]   
            await asyncio.sleep(0.5)      
            await message.answer(mess)

@router.message(Search.index)
async def comm_search (message: Message, state: FSMContext):
    mess_text = False
    for work_sheet in inv_book.worksheets:
        for row in work_sheet.iter_rows(min_row=1, max_row=work_sheet.max_row, min_col=1, max_col=work_sheet.max_column):
            for cell in row:
                if str(cell.value) == str(message.text):
                    mess_text = True
                    sheet = inv_book[work_sheet.title]
                    text = f"📃 [{work_sheet.title}] (Стр.: {cell.row} | Столб.: {cell.column})\n\n🔢 № {sheet['A' + str(cell.row)].value}\n📅 Дата ввода: {sheet['B' + str(cell.row)].value}\n✏️ Сер. номер: {sheet['C' + str(cell.row)].value}\n📌 Инв. номер: {sheet['D' + str(cell.row)].value}\n📝 Наименование: {sheet['E' + str(cell.row)].value}\n🚪 Местоположение: {sheet['F' + str(cell.row)].value}\n\n"
                    await asyncio.sleep(0.5)       
                    await message.answer(text)
    if mess_text == False:
        await message.answer("❌ Данного оборудования не найдено!")

@router.message(F.text.lower() == "📊 общая статистика")
async def comm_stats (message: Message):
    countsb = 0
    countmon = 0
    for work_sheet in inv_book.worksheets:  
        for row in work_sheet.iter_rows(min_row=1, max_row=work_sheet.max_row):
            for cell in row:
                if cell.value == "Системный блок": 
                    countsb += 1   
                elif cell.value == "Монитор": 
                    countmon += 1  
    await message.answer(f"📊 Общая статистика\n\n🖥 Кол-во мониторов: {countsb} шт.\n💻 Кол-во системных блоков: {countmon} шт.")
                    
                

  




    
